import platform
import sqlite3
from collections import namedtuple
from hashlib import sha256
from itertools import product
from pathlib import Path
from uuid import uuid4

import exif
import yaml
from openpyxl import Workbook, load_workbook


def named_tuples(cursor):
    Record = namedtuple("Record", [i[0] for i in cursor.description])
    return [Record._make(i) for i in cursor]


def create_data_file(path: str) -> None:
    """Read monexif_fields.yml and create a new workbook, *OVERWRITING* existing."""
    fields = yaml.safe_load(Path(__file__).with_name("monexif_fields.yml").open())
    wb = Workbook()
    ws = wb.active
    ws.title = "ImageData"
    for field_i, field in enumerate(fields["fields"]):
        ws.cell(0 + 1, field_i + 1, field)
    wb.save(path)


def insert_row(con, fields: list[str], row: list = None) -> None:
    if row is None:  # dict input
        row = list(fields.values())
        fields = list(fields)
    sql = "insert into imgdata (%s)" % ",".join(fields) + "values (%s)" % ",".join(
        "?" * len(fields)
    )
    con.execute(sql, row)


def field_defs():
    defs = yaml.safe_load(Path(__file__).with_name("monexif_fields.yml").open())
    for k, v in defs["fields"].items():
        v["name"] = k
    return defs


def xlsx_to_sqlite(xlsx_path: str, sqlite_path) -> object:
    """Convert .xlsx to .db, *OVERWRITING* existing imgdata table."""
    field_type = {k: v["type"] for k, v in field_defs()["fields"].items()}
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows = iter(ws)
    fields = [cell.value for cell in next(rows)]
    sql = [field + " " + field_type.get(field, "text") for field in fields]
    sql += [
        field + " " + field_type.get(field, "text")
        for field in field_type if field not in fields
    ]
    sql = "create table imgdata (\n" + ",\n".join(sql) + "\n)"
    con = sqlite3.connect(sqlite_path)
    con.execute("drop table if exists imgdata")
    con.execute(sql)
    for row in rows:
        insert_row(con, fields, [i.value for i in row])
    return con


def sqlite_to_xlsx(con, xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows = iter(ws)
    fields = [cell.value for cell in next(rows)]
    ws.delete_rows(2, ws.max_row)
    sql = "select " + ",".join(fields) + " from imgdata"
    res = con.execute(sql)
    for row in res:
        ws.append(row)
    wb.save(xlsx_path)


def add_images(con, basedir: str, paths: list[str]) -> int:
    for path in paths:
        res = con.execute("select * from imgdata where image_path = ?", [path])
        if list(res):
            print(f"Skipping existing {path}.")
            continue
        fullpath = Path(basedir) / path
        exif = read_exif(fullpath)
        data = fullpath.read_bytes()
        row = dict(
            image_name=fullpath.name,
            image_path=path,
            image_bytes=len(data),
            image_time=exif["datetime_original"].replace(":", "/", 2),
            image_w=exif["pixel_x_dimension"],
            image_h=exif["pixel_y_dimension"],
            image_hash=sha256(data).hexdigest(),
            observation_id=uuid4().hex,
            group_id=uuid4().hex,
            group_number=1,
        )
        insert_row(con, row)


def image_list(path: str) -> list[str]:
    """List of images from path *and its subfolders*."""

    if platform.system() == "Windows":
        transforms = [lambda x: x]  # case insensitive filesystem
    else:
        transforms = (str.upper, str.lower, str.title)

    image_paths = []
    #  Iterate .jpg, .JPEG, .Png etc.
    for transform, extension in product(transforms, ("jpg", "png", "jpeg", "gif")):
        ext = transform(extension)
        image_paths.extend(
            [
                str(i.resolve(strict=True).relative_to(path))
                for i in Path(path).rglob(f"*.{ext}")
            ]
        )
    return image_paths


def read_exif(path: str) -> dict:
    img = exif.Image(Path(path).open("rb"))
    result = {}
    for key in img.list_all():
        try:
            result[key] = getattr(img, key)
        except Exception:  # a zoo of different exceptions possible
            continue
    # import pprint
    # pprint.pprint(result)
    return result


def image_time_filename(exif: dict) -> str:
    """FIXME: assumes .jpg"""
    return exif["datetime_original"].replace(":", "").replace(" ", "_") + ".jpg"


def new_image_names(
    basedir: str, paths: list[str], do_renames: bool = False
) -> list[(str, str)]:
    renames = []
    for img_path in paths:
        exif = read_exif(Path(basedir) / img_path)
        name = image_time_filename(exif)
        if Path(img_path).name != name:
            new_path = Path(img_path).with_name(name)
            renames.append((img_path, new_path))
            if do_renames:
                (Path(basedir) / img_path).rename(Path(basedir) / new_path)
    return renames


def check_new(con: object, img_path: str):
    img_paths = set(image_list(img_path))
    img_recs = set(map(lambda x: x[0], con.execute("select image_path from imgdata")))
    print("")
    print(f"{len(img_paths)} images in folder.")
    print(f"{len(img_recs)} records in data.")
    print(f"{len(img_paths-img_recs)} in folder only.")
    print(f"{len(img_recs-img_paths)} in data only.")
    print(f"{len(img_recs&img_paths)} in both.")


def load_new(con: object, img_path: str):
    img_paths = set(image_list(img_path))
    img_recs = set(map(lambda x: x[0], con.execute("select image_path from imgdata")))
    new = img_paths - img_recs
    print(f"Adding {len(img_paths-img_recs)} images.")
    add_images(con, img_path, new)


def unset_related(con: object, obs_id: str) -> None:
    cur = con.cursor()
    cur.execute(
        "update imgdata set group_id=? where observation_id=?", [uuid4().hex, obs_id]
    )


def set_related(con: object, obs_id0: str, obs_id1: str) -> None:
    """Move obs_id0 into the same group as obs_id1"""
    cur = con.cursor()
    cur.execute(
        "select * from imgdata where observation_id in (?, ?)",
        [obs_id0, obs_id1],
    )
    obs = named_tuples(cur)
    # *** need to maintain ordering ***
    if obs[0].observation_id != obs_id0:
        obs.reverse()

    cur.execute(
        "update imgdata set group_id=? where observation_id=?",
        [
            obs[1].group_id,
            obs[0].observation_id,
        ],
    )
    for from_, to in (0, 1), (1, 0):
        updates = {}
        for field_name, field in field_defs()["fields"].items():
            if field.get("copy") and getattr(obs[to], field_name, None) is None:
                updates[field_name] = getattr(obs[from_], field_name, None)
        if updates:
            q = ["update imgdata set"]
            fields = ", ".join(f"{i}=?" for i in updates)
            q.append(fields)
            q.append("where observation_id=?")
            q = " ".join(q)
            print(q)
            values = list(updates.values()) + [obs[to].observation_id]
            print(values)
            con.execute(q, values)


if __name__ == "__main__":
    # print(image_list("pics"))
    # create_data_file("test.xlsx")
    con = xlsx_to_sqlite("test.xlsx", ":memory:")
    add_images(con, image_list("pics"))
    sqlite_to_xlsx(con, "test.xlsx")
    # print(new_image_names(image_list("pics")))
